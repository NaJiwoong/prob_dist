import random
import csv
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.utils.cell import get_column_letter

def Transpose(A):                           # Get matrix A as an input and return a transpose matrix of A
    row = len(A)
    col = len(A[0])
    A_T = list()
    for j in range(col):
        temp = list()
        for i in range(row):
            temp.append(A[i][j])
        A_T.append(temp)
    return A_T

class Dist:
    def __init__(self):
        self.prob_list = list()                 # List that has problem lists of chapters for element
        self.dist_num = list()
        self.chapter_name = list()
        self.chapter_num = 0                    # The number of chapters
        self.prob_per_chap = list()             # The max number of problems per chapters
        self.period = 0                         # Period
        self.prob_dist = list()                 # Final result that will be written to the csv file.
        self.dist_trans = list()                # Final result that is transposed
        self.max_num = 0
        self.index_chap_name = list()

    def parse_prob(self):
        prob_file = open("./problem_list.txt", 'rt', encoding='utf-8')
        sline = prob_file.readline()
        ch = sline.split(':')
        self.period = int(ch[1])
        while True:
            sline = prob_file.readline()
            if not sline:
                break
            line = sline.split(':')
            self.chapter_name.append(line[0].strip())
            line = line[1].split(',')
            self.update_max(len(line))
            for i in range(len(line)):
                line[i] = line[i].strip()
            self.prob_list.append(line)
            self.chapter_num += 1

    def update_max(self, prob_num):
        if self.max_num < prob_num:
            self.max_num = prob_num

    def get_period(self):
        while True:
            try:
                self.period = int(input("Enter period you want> "))
                break
            except ValueError:
                print("Only integer!")

    def shuffle(self):
        for chap_prob in self.prob_list:
            random.shuffle(chap_prob)

    def get_sequence(self):
        for chap_prob in self.prob_list:
            num_prob = len(chap_prob) // self.period
            num_rest = len(chap_prob) % self.period
            temp_list = [num_prob for i in range(self.period)]
            for i in range(num_rest):
                temp_list[i] += 1
            random.shuffle(temp_list)
            self.dist_num.append(temp_list)
            if num_rest > 0:
                self.prob_per_chap.append(num_prob+1)
            else:
                self.prob_per_chap.append(num_prob)

    def make_dist(self):
        for date in range(self.period):
            one_day = list()
            one_day.append("Day {0}".format(date+1))
            for ch_num in range(self.chapter_num):
                one_day.append(self.chapter_name[ch_num])
                prob_num = self.dist_num[ch_num][date]
                for prob in range(prob_num):
                    one_day.append(str(self.prob_list[ch_num].pop()))
                if prob_num < self.prob_per_chap[ch_num]:
                    one_day.append("")
            self.prob_dist.append(one_day)
        self.dist_trans = Transpose(self.prob_dist)
        for chapt in self.chapter_name:
            self.index_chap_name.append(self.prob_dist[0].index(chapt))

    def make_csv(self):
        f = open('problem.csv', 'w', encoding='utf-8-sig', newline='')
        wr = csv.writer(f)
        for list1 in self.dist_trans:
            wr.writerow(list1)
        f.close()

    def make_excel_file(self):
        filepath = "./problem.xlsx"
        wb = openpyxl.Workbook()

        sheet = wb.active

        for chapt_idx in self.index_chap_name:
            for j in range(self.period):
                sheet.merge_cells(start_row=chapt_idx+1, start_column=5*j+1, end_row=chapt_idx+1, end_column=5*j+5)

        for day in range(self.period):
            col_pivot = 5*day
            daily_list = self.prob_dist[day]

            sheet.cell(row=1, column=col_pivot+2).value = "주제"
            sheet.cell(row=1, column=col_pivot+2).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row=1, column=col_pivot+2).font = Font(color='006100')
            sheet.cell(row=1, column=col_pivot+2).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            for i in range(3):
                sheet.cell(row=1, column=col_pivot+3+i).value = "{0}".format(i+1)
                sheet.cell(row=1, column=col_pivot+3+i).alignment = Alignment(horizontal='center', vertical='center')
                sheet.cell(row=1, column=col_pivot+3+i).font = Font(color='006100')
                sheet.cell(row=1, column=col_pivot+3+i).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            for idx in range(len(daily_list)):
                sheet.cell(row=idx+1, column=col_pivot+1).value = daily_list[idx]
                sheet.cell(row=idx+1, column=col_pivot+1).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
                sheet.cell(row=idx+1, column=col_pivot+1).font = Font(color='006100')
                sheet.cell(row=idx+1, column=col_pivot+1).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

        for colidx in range(5*self.period):
            for idx in range(len(self.prob_dist[day])):
                sheet.cell(row=idx+1, column=colidx+1).border = Border(Side('thick'),Side('thick'),Side('thick'),Side('thick'))

        for idx in range(self.period):
            colidx = 5*idx
            sheet.column_dimensions[get_column_letter(colidx+1)].width = 10
            sheet.column_dimensions[get_column_letter(colidx+2)].width = 25
            sheet.column_dimensions[get_column_letter(colidx+3)].width = 3.5
            sheet.column_dimensions[get_column_letter(colidx+4)].width = 3.5
            sheet.column_dimensions[get_column_letter(colidx+5)].width = 3.5

        try:
            wb.save(filepath)
        except:
            print("엑셀 파일이 열려있지 않은지 확인하고 다시 시도해주세요.")

def main():
    my_dist = Dist()
    my_dist.parse_prob()
    my_dist.get_sequence()
    my_dist.shuffle()
    my_dist.make_dist()
    my_dist.make_csv()
    my_dist.make_excel_file()


if __name__ == "__main__":
    main()






